from email import message
from re import M
from time import time
import requests
import json
import datetime
import os
import sched, time
import logging
import smtplib

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta

##Dependencies##

#requests
#openpyxl
#googleapiclient  pip install --upgrade google-api-python-client google-auth-httplib2 google-auth-oauthlib

# GOOGLE AUTH 
import os.path
from googleapiclient.discovery import build
from google.oauth2 import service_account

SERVICE_ACCOUNT_FILE = 'keys.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

creds = None
creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

SAMPLE_SPREADSHEET_ID = '1E5-AezYFSapijKRCwsLUfD16qsq2HTh3f3s-pgGFzNo'



service = build('sheets', 'v4', credentials=creds)

# Call the Sheets API
sheet = service.spreadsheets()

#Ende

#Log
logging.basicConfig(filename="logCraw.log", level=logging.INFO, format='%(asctime)s - %(message)s')

datum = (datetime.now() - timedelta(1))
datum_weekday = datum.weekday()

if(datum_weekday < 5): 
    tradeDate = datum.strftime('%Y%m%d')
elif(datum_weekday == 5):
    tradeDate = (datum - timedelta(1)).strftime('%Y%m%d')
elif(datum_weekday == 6):
    tradeDate = (datum - timedelta(2)).strftime('%Y%m%d')


#Get trade Date from Website

headers = {
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate",
    "User-Agent": "Mozilla/5.0"
}

logging.info("START craw.py")
s = sched.scheduler(time.time, time.sleep)


def checkTradeDate(tradeDate_data, tradeDate, url_id, params):

    #While Array Len == 0 -> tradeDate nicht richtig
    #Wihile Array Len > 0 -> tradeDate richtig

    while(len(tradeDate_data["monthData"]) == 0):
        tradeDateInt = int(tradeDate) - 1
        tradeDate = str(tradeDateInt)
        
        newUrl = "https://www.cmegroup.com/CmeWS/mvc/Volume/Details/F/"+ url_id +"/"+ tradeDate + "/P"
        try: #CME-Group
            response = requests.get(newUrl, params=params, headers=headers) #URL
        except requests.exceptions.Timeout:
            logging.exception("Exception occured (timeout) - Connect to CME")
            print("time-out")
        except requests.exceptions.ConnectionError:
            logging.exception("Exception occured (conn err) - Connect to CME")
            print('Connection Error')
                

        response.raise_for_status()

        tradeDate_data = response.json() #Data von Url Json

    
        
    return tradeDate_data
    
def printProgressBar(progress):

    print("[", end="")

    for i in range(0 , 100):
        if(i < progress): print("=", end="")
        if(i > progress): print("-" , end="")

    
    print("]")

def sendmail():
    host = "smtp.freesmtpservers.com"
    server = smtplib.SMTP(host)
    FROM = "tester@test.de"
    TO = "kalhornpaul@gmail.com"
    MSG = "Subject: Test email python\n\nBody of your message!"
    server.sendmail(FROM, TO, MSG)
    
def main(sc):

    logging.info("Start Process CRAW")

    wb = load_workbook('Daten.xlsx') 

    ws = wb.active

    dataName_array = ["Currencies", "Energies" , "Equities" , "Financials" , "Grains" , "Meats" , "Metals" , "Softs"]

    b = 0

    progress = 0.0

    isCme = False


    cord_col_a = 1
    cord_col_b = 2 

    print(            """\n                                                                     
                                        .......                                               
                                :::::=@@@@@@@-::::.                                         
                            .::-%@@%%%*******#%@@@#::::::::::::::::::.                      
                        .---*%%%******************#%%%%%%%%####%%%%%%*-------               
                        .=*%%%#******#%%%#*****************************#%%%%%%%===-           
                    =+%#*********%#*#@%%#***********************************%%%#=-         
                    ++%#***********@@+*#%@%****************####***##########*****#%#+-       
                    @@*************%%@@%%%#************#####%%%########%%%%@#******#%#+:     
                **@@*************##%%%%#*************#####%%%%%%###*###%%%%%*******%@=     
                +#****%%%%@@@%#******###***********#####***##%%%##%%%###%%%%@@***###*####:   
            +#*+====+++++++#%@@##*********+++*****##%######%###%%%%%%%%%%%%###*#%#*##@@-   
            =#*+====*****#%%%@@@@@@####****+***###****##%%%%###%%%%%%%%%%%%%%*%#+#%#*#%##%%. 
        =%*+==+*****@@@#+++++=--=@@@@%##%###%%%%####***####%%%%%%%%%%%%%%%%##%@%*###*#%@@: 
        -@#===+*****@%-=+-.:---%@@@**@@%###*****%#%%%%%%%#%%%%%%%%#%%%%%%%##%%%%@@*#%#*#%@@: 
        -@#=+***%@@@--:-+=-*@@@+=**@@%%%%%%%#░█████╗░██████╗░░█████╗░░██╗░░░░░░░██╗░░░*#%#*@@
    :@%=+***%@+===++**@@@*==+****@@%%%%%%%%██╔══██╗██╔══██╗██╔══██╗░██║░░██╗░░██║░░░%%%##@@
    :-%#++*#%*+::--=+%#++++*****@@%%%%%%%%%%██║░░╚═╝██████╔╝███████║░╚██╗████╗██╔╝░░░%%%%%@@
    .@%=+##%+:=+=+%%%#++****#########%%%%#%%%██║░░██╗██╔══██╗██╔══██║░░████╔═████║░░░░%%%%%@@
    --%#**%*=-:+*#%++++**###****.    .***#%%%%╚█████╔╝██║░░██║██║░░██║░░╚██╔╝░╚██╔╝░██╗@%%%%@@
    @@=+@#===+###++****##**+:            =@%##░╚════╝░╚═╝░░╚═╝╚═╝░░╚═╝░░░╚═╝░░░╚═╝░░╚═╝@%%%%@@
    @@###+:=**#*+**####+-                :+*##%%%%%%%%%%%%%%%####%%%%%%%%%%%%###%%%%%%%@%%%%@@
    ===-:-**#*+#%%%===-                    +@%#%%%#%%%%%%%#%%%%%%%%#%%%%%%%###%%%%%%%@@%@@@@+=
    .***##%%+---                        :-#%##%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%@@@@%%%@@+-  
    :@@%+---.                             :-#%%%%%%%%%%%%%%%#%%%%%%%%%%%%%%%@@@@%%%%%@@-   
                                                @@##%%@@@@@@@@@@@@@@@@@@@@%%%%%%@@%%%%%@@-   
                                                @@##%%@*.*@%%%%%#%%%%%%%@@%%%%@@%%%%%%%%%%%. 
                                            *%@@%%@*     #@#*##%%%%%%%%%@@@%%##%%%%%%%%@@: 
                                            *@**@#         #@*#%%%%#%%%%@@*=%#*%%%@@@%%@@: 
                                            #@+=@@           #@###%%###%%%@@**@@%%%%@+ +@@@. 
                                            .*@#*%#             @@*#@@%####@@+=@@%%%%@+       
                                        +@*+%@               ####@* -###@@+=@@##%%@+       
                                        .:*%####                 ###+   =#@@+=@% .###=       
                                        =@#+%@.                           #@++@@             
                                    .-+###**.                           #@**@@             
                                    =+#*+#@-                             #@**@%             
                                =+++#*++*%@-                         :+++*#**@%             
                                =+**###*+*%##*+.                     .+*#@@+=####+=           
                                %@==%%@%+#@#*%@.                   .***#%@@*+@@**@#           
                                %@%#-=@@%*-*%@@.                   :@@%*-#@@%--%%@#           
                                ::::  :::. .:::                     :::. .:::  :::.           
                                    """ )

    #sendmail()

    inc = 4
    ch = 'B'
    #For Schleife -> Gehe durch Datensätze (in Data ordner)
    for b in range (0, 8):

        inc = 4 
    
        if b >= 1: 
            cord_col_a = cord_col_a + 3
            cord_col_b = cord_col_b + 3

        with open("data/" + dataName_array[b] + ".json") as f:
            info_data = json.load(f)

        ws[get_column_letter(cord_col_a) + "2" ] = dataName_array[b]

        cord_a = 4
        cord_b = 5

        print("Fetching: " + dataName_array[b] + ".json")

        printProgressBar(progress)

        print("Progress: " + str(progress) + "%\n")

        

        for l in range(0 , len(info_data["infoData"])):

            rangeS = "Kontraktvolumen!" + str(ch) + str(inc) #Name der Tabelle auf GoogleSheets! 
            
            progress = round(progress + (12.5/len(info_data["infoData"])), 2)

            if(info_data["infoData"][l]["from"] == "theice"): #Ist TheIce url?
                
                isCme = False

                params_ice = {
                    "getContractsAsJson": "",
                    "productId": info_data["infoData"][l]["url-id"], 
                    "hubId": info_data["infoData"][l]["hub-id"], 
                }   

                url_ice = "https://www.theice.com/marketdata/DelayedMarkets.shtml?"

                try: #TheIce
                    response_Ice = requests.get(url_ice, params=params_ice, headers=headers)
                except requests.exceptions.RequestException as e1:
                    logging.exception("Exception occured - Connect to TheICE")
                    raise SystemExit(e1)

                response_Ice.raise_for_status()

                data_ice = response_Ice.json() #Data von Url Json

            if(info_data["infoData"][l]["from"] == "cme"): #Ist cme url?

                
                params = {
                    "tradeDate": tradeDate, #wie verändert sich das trade datum?
                    "pageSize": "50",
                    "_": "1620683546888"
                }

                url_id = (info_data["infoData"][l]["url-id"])

                url = "https://www.cmegroup.com/CmeWS/mvc/Volume/Details/F/"+ url_id +"/" + tradeDate + "/P"
                print(url)
                try: #CME-Group
                    response = requests.get(url, params=params, headers=headers) #URL
                except requests.exceptions.Timeout:
                    logging.exception("Exception occured (timeout) - Connect to CME")
                    print("time-out")
                except requests.exceptions.ConnectionError:
                    logging.exception("Exception occured (conn err) - Connect to CME")
                    print('Connection Error')
                

                response.raise_for_status()

                data_Cme_NC = response.json() #Data von Url Json

                data_Cme = checkTradeDate(data_Cme_NC, tradeDate, url_id, params)

                isCme = True

                


            ws[get_column_letter(cord_col_a) + str(cord_a)] = info_data["infoData"][l]["name"]
            ws[get_column_letter(cord_col_a) + str(cord_b)] = "Monat:"
            ws[get_column_letter(cord_col_b) + str(cord_b)] = "Total Volume:"
            name = info_data["infoData"][l]["name"]
            i = 0 #für schleife monate reset
            aoa = [[name], ["MONAT", "TOTAL"]]
        
            cord_b = cord_b + 1

            while i < 3:

                if(isCme == True):
                    if i == len(data_Cme["monthData"]): #Exit wenn nicht mehr monate # Dann im array leer!
                        cord_b = cord_b + 1
                        break
    
                    ws[get_column_letter(cord_col_a) + str(cord_b)] = data_Cme["monthData"][i]["month"]
                    ws[get_column_letter(cord_col_b) + str(cord_b)] = data_Cme["monthData"][i]["totalVolume"]

                    month = data_Cme["monthData"][i]["month"]
                    totalVolume = data_Cme["monthData"][i]["totalVolume"]
                    arrayMonat = [month, totalVolume]
                    aoa.append(arrayMonat)

                    
                if(isCme == False):

                    if i == len(data_ice): #Bug fix "Out of Range z.190"
                        cord_b = cord_b + 1
                        break
                    
                    #bug: Keine Zeit gefunden, an was liegt es Test: 00:56 Uhr vllt deswegen? -> ws[x] lässt nicht schreiben nur lesen??
                    #ws[get_column_letter(cord_col_a) + str(cord_b)] = data_ice[i]["marketStrip"] 
                    #ws[get_column_letter(cord_col_b) + str(cord_b)] = data_ice[i]["volume"]

                    month = data_ice[i]["marketStrip"]
                    totalVolume = data_ice[i]["volume"]
                    arrayMonat = [month, totalVolume]
                    aoa.append(arrayMonat)


                i = i + 1 #Increment
            
                cord_b = cord_b + 1

            inc = inc + 6

            #Write aoa to GoogleSheet
            try:
                request = sheet.values().update(spreadsheetId=SAMPLE_SPREADSHEET_ID, 
                                    range=rangeS ,valueInputOption="USER_ENTERED", body={"values": aoa}).execute()
            except: 
                logging.exception("Connection error - could not be pushed to sheet")
            
            #Abstand zwischen neuen Datensätzen:
            cord_b = cord_b + 2
            cord_a = cord_a + 6

            i = 0 #reset für nächsten durchlauf

        ch = chr(ord(ch) + 3) #Increment in GoogleSheet(-> Jeder Datensatz soll nebeneinander stehen B wird zu B+3=E)
        wb.save("Daten.xlsx")

    print(""" \n 
    ███████╗██╗███╗░░██╗██╗░██████╗██╗░░██╗██╗
    ██╔════╝██║████╗░██║██║██╔════╝██║░░██║██║
    █████╗░░██║██╔██╗██║██║╚█████╗░███████║██║
    ██╔══╝░░██║██║╚████║██║░╚═══██╗██╔══██║╚═╝
    ██║░░░░░██║██║░╚███║██║██████╔╝██║░░██║██╗
    ╚═╝░░░░░╚═╝╚═╝░░╚══╝╚═╝╚═════╝░╚═╝░░╚═╝╚═╝
    """)
    logging.info("Finished Process CRAW")
    s.enter(300, 1, main, (sc,)) #Repeat every 5min
    
##RUN EACH 60 SEC ###############
s.enter(5, 1, main, (s,))
s.run()
#################################

#Notifiyer wenn Crash
#Immer aktuellsten Datensatz haben nicht ab 0 Uhr z.B den noch nicht aktuellen
#Log wird aktuell auf Vserver nicht geschrieben